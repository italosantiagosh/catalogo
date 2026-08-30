"""
Gera a planilha de importacao em massa da Shopee (mesmo modelo baixado
direto no Seller Center -- "Modelo de criacao em massa", aba "Modelo")
a partir do catalogo real do site (data/produtos.json), preenchendo as
linhas de dados a partir da linha 7 do arquivo real enviado pelo
usuario (linhas 1-6 sao cabecalho/instrucao da propria Shopee, mantidas
intactas).

Uso: python3 scripts/gerar_planilha_shopee.py <template_baixado.xlsx> [saida.xlsx]

Por que precisa do template real como entrada (e nao gera do zero):
o arquivo que a Shopee exporta tem varias abas auxiliares (instrucoes,
tabela de impostos, lista de unidades de medida) e a ordem exata das
colunas pode mudar entre categorias/exportacoes -- ler o cabecalho de
verdade (linha 1, chaves internas tipo "ps_product_name") em vez de
hardcodar letra de coluna evita gerar planilha com coluna errada se a
Shopee mudar algo.

Bug conhecido do proprio export da Shopee: o XML da planilha vem com
`activePane="bottom_left"` (underscore), valor invalido no padrao OOXML
-- isso faz o Excel/LibreOffice/openpyxl recusarem abrir o arquivo.
Esse script conserta isso automaticamente (troca por "bottomLeft") antes
de abrir, tanto no arquivo de entrada quanto ao salvar a saida.

ESTRUTURA (ver conversa -- diferente da planilha da Tiny de proposito):
  - 1 anuncio por SANTO + FORMATO (nao mais por santo+modelo) -- "Modelo"
    agora e´ a propria Variacao 1, pra nao duplicar anuncio por modelo.
  - Medalha:   Variacao 1 = Modelo   | Variacao 2 = Tamanho (12mm/16mm)
  - Entremeio: Variacao 1 = Modelo   | Variacao 2 = Cor (Ouro Velho/Prata)
  - Chaveiro:  Variacao 1 = Modelo   | sem Variacao 2 (chaveiro nao varia)
  - Preco fixo por formato, DIFERENTE do preco da Tiny de proposito (ver
    conversa): R$15,00 medalha/entremeio, R$25,00 chaveiro.
  - SKU por variacao reaproveita o MESMO esquema ja confirmado em
    services/tiny.py::_codigo_estoque_tiny (MED-/ENT-/CHAV-<id>-M<modelo>-...),
    pra ficar facil bater com o cadastro da Tiny se precisar conferir.

LIMITACOES ASSUMIDAS (avisadas no console ao rodar -- conferir antes de
subir pra Shopee):
  - Categoria (coluna A): deixada em branco -- nao temos o ID numerico
    da categoria da Shopee. Usuario pediu "Hobbies e Colecoes >
    Souvenires > Outros" (medalha/entremeio) e "> Chaveiros" (chaveiro);
    a Shopee sugere automaticamente quando fica em branco, mas confirmar
    antes de publicar.
  - Estoque: preenchido com um numero fixo (ver ESTOQUE_PLACEHOLDER)
    so´ porque o campo e´ obrigatorio pra publicar -- produto e´ feito
    sob encomenda, sem estoque fisico real Pra ajustar o valor.
  - Campos fiscais que dependem do regime tributario da empresa (CFOP
    mesmo/outro estado, CSOSN, CEST, % total de tributos, Tipo de
    Operacao, CST PIS/Cofins) ficam em BRANCO de proposito -- nunca
    inventar codigo fiscal. NCM, Origem e Unidade de medida SAO
    preenchidos (mesmos dados ja confirmados no script da Tiny).
  - GTIN/EAN: em branco (produto sem codigo de barras cadastrado).
"""

from __future__ import annotations

import shutil
import sys
import tempfile
import zipfile
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from gerar_planilha_tiny import (  # noqa: E402
    DOMINIO,
    NCM,
    ORIGEM,
    PESO_KG_POR_CHAVE,
    _DESCRICOES_POR_FORMATO,
)

BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "data"

LARGURA_CM = 11
ALTURA_CM = 3
COMPRIMENTO_CM = 16

PRECO_SHOPEE_POR_FORMATO = {"medalha": 15.00, "entremeio": 15.00, "chaveiro": 25.00}
ESTOQUE_PLACEHOLDER = 999

# Prazo de Postagem para Encomenda -- usuario pediu 3 dias (o mais
# rapido que o real prazo de producao, config.py::PRODUCAO_DIAS_UTEIS=5,
# permitiria arredondar pra baixo) "ou o minimo, se for mais que isso".
# Conferido na aba oficial "Intervalo do PP para Encomenda" do proprio
# template da Shopee: as duas categorias abaixo aceitam faixa "3 - 15"
# dias, entao 3 ja´ e´ o minimo permitido -- nao precisa subir.
PRAZO_POSTAGEM_DIAS = 3

# TESTE (ver conversa): categoria em branco de proposito, pra isolar se
# 101399/101396 e´ que rejeitava produto sob encomenda (Prazo de
# Postagem preenchido). Se der certo com categoria em branco, o
# problema era a combinacao categoria+encomenda; se continuar falhando
# do mesmo jeito, o proximo teste e´ o oposto (categoria preenchida,
# sem prazo de encomenda).
CATEGORIA_ID_POR_FORMATO = {"medalha": "", "entremeio": "", "chaveiro": ""}
CATEGORIA_TEXTO_POR_FORMATO = {
    "medalha": "101399 - Hobbies e Coleções/Souvenirs/Outros",
    "entremeio": "101399 - Hobbies e Coleções/Souvenirs/Outros",
    "chaveiro": "101396 - Hobbies e Coleções/Souvenirs/Chaveiros",
}

ORIGEM_TEXTO = {0: "0 - Nacional, exceto as indicadas nos códigos 3, 4, 5 e 8"}[ORIGEM]
UNIDADE_MEDIDA_TEXTO = "UN (UNIDADE)"


def _consertar_e_abrir(caminho: Path) -> openpyxl.Workbook:
    """Corrige o `activePane="bottom_left"` invalido (ver docstring do
    modulo) numa copia temporaria e abre com openpyxl."""
    tmp_dir = Path(tempfile.mkdtemp(prefix="shopee_fix_"))
    extraido = tmp_dir / "extraido"
    with zipfile.ZipFile(caminho) as zf:
        zf.extractall(extraido)

    for sheet_xml in (extraido / "xl" / "worksheets").glob("sheet*.xml"):
        texto = sheet_xml.read_text(encoding="utf-8")
        if 'activePane="bottom_left"' in texto:
            sheet_xml.write_text(texto.replace('activePane="bottom_left"', 'activePane="bottomLeft"'), encoding="utf-8")

    consertado = tmp_dir / "consertado.xlsx"
    with zipfile.ZipFile(consertado, "w", zipfile.ZIP_DEFLATED) as zf:
        for arquivo in extraido.rglob("*"):
            if arquivo.is_file():
                zf.write(arquivo, arquivo.relative_to(extraido))

    return openpyxl.load_workbook(consertado)


def _mapa_de_colunas(ws) -> dict[str, int]:
    """Le a linha 1 (chaves internas tipo "ps_product_name|1|0" -- a
    Shopee sufixa cada chave com "|obrigatorio|algumOutroFlag", ver
    conversa) e devolve {chave_sem_sufixo: indice_da_coluna}. Nunca
    hardcoda letra de coluna, pra resistir a planilha da Shopee mudar a
    ordem entre exportacoes."""
    mapa: dict[str, int] = {}
    for celula in ws[1]:
        if celula.value:
            chave = str(celula.value).split("|", 1)[0]
            mapa[chave] = celula.column
    return mapa


def _url_imagem(caminho_relativo: str) -> str:
    return f"{DOMINIO}/static/{caminho_relativo}"


def _linhas_do_produto(produto: dict, formato: str) -> list[dict[str, object]]:
    """Monta as linhas de UM anuncio (santo + formato), com Modelo como
    Variacao 1 e Tamanho/Cor como Variacao 2 (exceto chaveiro, sem
    Variacao 2) -- ver docstring do modulo."""
    produto_id = produto["id"]
    nome_santo = produto["nome"]
    prefixo = {"medalha": "MED", "entremeio": "ENT", "chaveiro": "CHAV"}[formato]
    sku_pai = f"{prefixo}-{produto_id.upper()}"
    preco = PRECO_SHOPEE_POR_FORMATO[formato]
    imagem_capa = _url_imagem(produto["modelos"][0]["imagem" if formato == "medalha" else (
        "imagem_entremeio_prata" if formato == "entremeio" else "imagem_chaveiro"
    )])

    base_comum = {
        "ps_category": CATEGORIA_ID_POR_FORMATO[formato],
        "ps_product_name": {
            "medalha": f"Medalha {nome_santo} - Aço Inox Resinado - Nove de Julho",
            "entremeio": f"Entremeio {nome_santo} para Terço - Nove de Julho",
            "chaveiro": f"Chaveiro {nome_santo} - Nove de Julho",
        }[formato],
        "ps_product_description": _DESCRICOES_POR_FORMATO[formato](nome_santo),
        "ps_sku_parent_short": sku_pai,
        "et_title_variation_integration_no": sku_pai,
        "ps_stock": ESTOQUE_PLACEHOLDER,
        "ps_item_cover_image": imagem_capa,
        # "Ativar"/"Off" -- CORRIGIDO (ver conversa): a linha 6 da
        # coluna descreve "Ativado/Desativado", mas essa e´ so a
        # instrucao em texto -- os valores de verdade usados na propria
        # aba "Fazer upload do exemplo" da Shopee (dados reais de
        # exemplo, nao descricao) sao "Ativar"/"Off". "Ativado" (usado
        # antes) tambem falhava, entao a descricao da linha 6 nao
        # reflete o token de verdade esperado. 90024 e´ "Retirada pelo
        # Comprador" -- usuario decidiu nao oferecer essa opcao (ver
        # conversa), entao fica Off de proposito (nao e´ erro, e´
        # escolha). 91003 e´ "Shopee Xpress", o unico canal que o
        # usuario realmente quer.
        "channel_id.90024": "Off",
        "channel_id.91003": "Ativar",
        "ps_product_pre_order_dts": PRAZO_POSTAGEM_DIAS,
        "ps_invoice_ncm": NCM.replace(".", ""),
        "ps_invoice_origin": ORIGEM_TEXTO,
        "ps_invoice_measure_unit": UNIDADE_MEDIDA_TEXTO,
        "ps_length": COMPRIMENTO_CM,
        "ps_width": LARGURA_CM,
        "ps_height": ALTURA_CM,
    }

    linhas: list[dict[str, object]] = []
    for modelo in produto["modelos"]:
        modelo_id = modelo["id"]
        modelo_nome = modelo["nome"]

        if formato == "medalha":
            imagem_variacao = _url_imagem(modelo["imagem"])
            for tamanho in modelo["tamanhos"]:
                linhas.append({
                    **base_comum,
                    "et_title_variation_1": "Modelo", "et_title_option_for_variation_1": modelo_nome,
                    "et_title_image_per_variation": imagem_variacao,
                    "et_title_variation_2": "Tamanho", "et_title_option_for_variation_2": tamanho,
                    "ps_price": preco, "ps_weight": PESO_KG_POR_CHAVE[tamanho],
                    "ps_sku_short": f"{sku_pai}-M{modelo_id}-{tamanho}",
                })
        elif formato == "entremeio":
            # A Shopee exige a MESMA imagem em todas as linhas que
            # compartilham a mesma opcao de Variacao 1 (aqui, Modelo) --
            # ver conversa/docstring do modulo. Por isso a imagem por
            # variacao usa sempre a foto do modelo (prata), nunca varia
            # por Cor (Variacao 2); a foto ouro velho entra como imagem
            # extra da galeria (S/ps_item_image_1) pra nao perder a
            # referencia visual da outra cor.
            imagem_variacao = _url_imagem(modelo["imagem_entremeio_prata"])
            imagem_extra = _url_imagem(modelo["imagem_entremeio_ouro_velho"])
            for cor in ("Ouro Velho", "Prata"):
                linhas.append({
                    **base_comum,
                    "et_title_variation_1": "Modelo", "et_title_option_for_variation_1": modelo_nome,
                    "et_title_image_per_variation": imagem_variacao,
                    "et_title_variation_2": "Cor", "et_title_option_for_variation_2": cor,
                    "ps_price": preco, "ps_weight": PESO_KG_POR_CHAVE["entremeio"],
                    "ps_sku_short": f"{sku_pai}-M{modelo_id}-{cor[:2].upper()}",
                    "ps_item_image_1": imagem_extra,
                })
        else:  # chaveiro -- so Variacao 1 (Modelo), sem Variacao 2
            linhas.append({
                **base_comum,
                "et_title_variation_1": "Modelo", "et_title_option_for_variation_1": modelo_nome,
                "et_title_image_per_variation": _url_imagem(modelo["imagem_chaveiro"]),
                "ps_price": preco, "ps_weight": PESO_KG_POR_CHAVE["chaveiro"],
                "ps_sku_short": f"{sku_pai}-M{modelo_id}",
            })
    return linhas


def gerar(template_path: Path, saida_path: Path) -> int:
    # Duas tentativas anteriores de "limpar" o arquivo pioraram as coisas
    # (ver historico do commit): apagar as abas auxiliares corrompeu o
    # arquivo (referencia de validacao pendurada), e montar um workbook
    # novo do zero passou no leitor xlsx generico mas a propria Shopee
    # rejeitou como "arquivo invalido" -- o uploader deles e´ mais
    # exigente que o padrao OOXML generico e espera o formato exato do
    # template que eles mesmos exportam. A 1a versao (so escrever os
    # dados dentro do template real, sem tocar em mais nada) foi a UNICA
    # que a Shopee aceitou estruturalmente (chegou a contar 405
    # produtos, so rejeitou por conteudo) -- entao e´ essa abordagem que
    # fica valendo: edita o workbook original de verdade, todas as abas
    # intactas, so escreve nas celulas da aba "Modelo".
    wb = _consertar_e_abrir(template_path)
    ws = wb["Modelo"]
    colunas = _mapa_de_colunas(ws)

    produtos = __import__("json").loads((DATA_DIR / "produtos.json").read_text(encoding="utf-8"))

    linha_atual = 7  # linhas 1-6 sao cabecalho/instrucao da Shopee -- nunca sobrescrever
    total = 0
    for produto in produtos:
        for formato in ("medalha", "entremeio", "chaveiro"):
            for valores in _linhas_do_produto(produto, formato):
                for chave, valor in valores.items():
                    if chave not in colunas:
                        continue
                    ws.cell(row=linha_atual, column=colunas[chave], value=valor)
                linha_atual += 1
                total += 1

    wb.save(saida_path)
    return total


def main() -> None:
    if len(sys.argv) < 2:
        print("Uso: python3 scripts/gerar_planilha_shopee.py <template_baixado.xlsx> [saida.xlsx]")
        sys.exit(1)
    template_path = Path(sys.argv[1])
    saida_path = Path(sys.argv[2]) if len(sys.argv) > 2 else BASE_DIR / "planilha_shopee_catalogo.xlsx"

    total = gerar(template_path, saida_path)
    print(f"{total} linhas geradas em {saida_path} (dentro do template original, todas as abas intactas)")
    print()
    cat_medalha = CATEGORIA_ID_POR_FORMATO["medalha"] or "(EM BRANCO -- teste, ver conversa)"
    cat_chaveiro = CATEGORIA_ID_POR_FORMATO["chaveiro"] or "(EM BRANCO -- teste, ver conversa)"
    print(f"Categoria: {cat_medalha} (medalha/entremeio), {cat_chaveiro} (chaveiro)")
    print(f"Prazo de Postagem para Encomenda: {PRAZO_POSTAGEM_DIAS} dias (minimo permitido pras duas categorias e´ 3)")
    print()
    print("CONFERIR ANTES DE SUBIR NA SHOPEE:")
    print(f"  - Estoque preenchido com {ESTOQUE_PLACEHOLDER} (placeholder, produto e´ sob encomenda) -- ajustar se quiser outro numero.")
    print("  - Campos fiscais que dependem do regime tributario (CFOP, CSOSN, CEST, % de tributos,")
    print("    Tipo de Operacao, CST PIS/Cofins) ficaram em branco de proposito -- preencher com o contador.")


if __name__ == "__main__":
    main()
